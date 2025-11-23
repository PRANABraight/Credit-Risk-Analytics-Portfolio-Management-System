"""
Automated Reporting and Data Refresh Workflows
Integrates with Excel and Google Sheets for automated KPI reporting
"""

import pandas as pd
import pymysql
from datetime import datetime, timedelta
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, LineChart, Reference
from scripts.config import config

class AutomatedReporter:
    """Automated reporting system for operational KPIs"""
    
    def __init__(self):
        self.conn = pymysql.connect(**config.get_db_config())
        self.output_dir = Path("reports")
        self.output_dir.mkdir(exist_ok=True)
    
    def generate_daily_report(self):
        """Generate daily operational report"""
        print("[INFO] Generating daily operational report...")
        
        # Fetch data
        daily_kpis = pd.read_sql("""
            SELECT * FROM vw_daily_kpis
            WHERE activity_date >= DATE_SUB(CURDATE(), INTERVAL 7 DAY)
            ORDER BY activity_date DESC
        """, self.conn)
        
        executive_kpis = pd.read_sql("SELECT * FROM vw_executive_kpis", self.conn)
        
        # Create Excel workbook
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = self.output_dir / f"daily_report_{timestamp}.xlsx"
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Executive Summary
            executive_kpis.T.to_excel(writer, sheet_name='Executive Summary')
            
            # Daily KPIs
            daily_kpis.to_excel(writer, sheet_name='Daily KPIs', index=False)
            
            # Format sheets
            self._format_executive_sheet(writer.book['Executive Summary'])
            self._format_daily_kpis_sheet(writer.book['Daily KPIs'], daily_kpis)
        
        print(f"[OK] Daily report saved: {output_file}")
        return output_file
    
    def generate_weekly_report(self):
        """Generate weekly performance report"""
        print("[INFO] Generating weekly performance report...")
        
        # Fetch data
        weekly_trends = pd.read_sql("""
            SELECT * FROM vw_weekly_performance_trends
            ORDER BY week_start DESC
            LIMIT 12
        """, self.conn)
        
        collection_efficiency = pd.read_sql("""
            SELECT * FROM vw_collection_efficiency
        """, self.conn)
        
        customer_engagement = pd.read_sql("""
            SELECT * FROM vw_customer_engagement
            ORDER BY engagement_score DESC
            LIMIT 100
        """, self.conn)
        
        # Create Excel workbook
        timestamp = datetime.now().strftime("%Y%m%d")
        output_file = self.output_dir / f"weekly_report_{timestamp}.xlsx"
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            weekly_trends.to_excel(writer, sheet_name='Weekly Trends', index=False)
            collection_efficiency.to_excel(writer, sheet_name='Collection Efficiency', index=False)
            customer_engagement.to_excel(writer, sheet_name='Customer Engagement', index=False)
            
            # Add charts
            self._add_weekly_trend_chart(writer.book['Weekly Trends'], weekly_trends)
        
        print(f"[OK] Weekly report saved: {output_file}")
        return output_file
    
    def export_to_google_sheets(self, sheet_name: str = "Credit Risk KPIs"):
        """
        Export KPIs to Google Sheets
        Requires google-auth and gspread packages
        """
        try:
            import gspread
            from google.oauth2.service_account import Credentials
            
            print("[INFO] Exporting to Google Sheets...")
            
            # Setup credentials (requires service account JSON)
            scopes = ['https://www.googleapis.com/auth/spreadsheets']
            creds = Credentials.from_service_account_file(
                'credentials/google_service_account.json',
                scopes=scopes
            )
            client = gspread.authorize(creds)
            
            # Get or create spreadsheet
            try:
                sheet = client.open(sheet_name)
            except gspread.SpreadsheetNotFound:
                sheet = client.create(sheet_name)
            
            # Export executive KPIs
            executive_data = pd.read_sql("SELECT * FROM vw_executive_kpis", self.conn)
            worksheet = sheet.worksheet('Executive KPIs') if 'Executive KPIs' in [ws.title for ws in sheet.worksheets()] else sheet.add_worksheet('Executive KPIs', rows=100, cols=20)
            worksheet.update([executive_data.columns.values.tolist()] + executive_data.values.tolist())
            
            # Export daily KPIs
            daily_data = pd.read_sql("""
                SELECT * FROM vw_daily_kpis
                WHERE activity_date >= DATE_SUB(CURDATE(), INTERVAL 30 DAY)
            """, self.conn)
            worksheet = sheet.worksheet('Daily KPIs') if 'Daily KPIs' in [ws.title for ws in sheet.worksheets()] else sheet.add_worksheet('Daily KPIs', rows=1000, cols=20)
            worksheet.update([daily_data.columns.values.tolist()] + daily_data.values.tolist())
            
            print(f"[OK] Data exported to Google Sheets: {sheet.url}")
            return sheet.url
            
        except ImportError:
            print("[WARNING] Google Sheets export requires: pip install gspread google-auth")
            return None
        except FileNotFoundError:
            print("[WARNING] Google service account credentials not found")
            print("         Place credentials at: credentials/google_service_account.json")
            return None
    
    def _format_executive_sheet(self, worksheet):
        """Format executive summary sheet"""
        # Header formatting
        for cell in worksheet['A']:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Adjust column width
        worksheet.column_dimensions['A'].width = 30
        worksheet.column_dimensions['B'].width = 20
    
    def _format_daily_kpis_sheet(self, worksheet, data):
        """Format daily KPIs sheet"""
        # Header row
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _add_weekly_trend_chart(self, worksheet, data):
        """Add trend chart to weekly report"""
        chart = LineChart()
        chart.title = "Weekly Performance Trends"
        chart.style = 10
        chart.y_axis.title = 'Payments'
        chart.x_axis.title = 'Week'
        
        data_ref = Reference(worksheet, min_col=2, min_row=1, max_row=len(data)+1, max_col=3)
        cats = Reference(worksheet, min_col=1, min_row=2, max_row=len(data)+1)
        
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats)
        
        worksheet.add_chart(chart, "F2")
    
    def close(self):
        """Close database connection"""
        if self.conn:
            self.conn.close()


def schedule_daily_report():
    """Scheduled task for daily report generation"""
    reporter = AutomatedReporter()
    try:
        reporter.generate_daily_report()
        # Optionally export to Google Sheets
        # reporter.export_to_google_sheets()
    finally:
        reporter.close()


def schedule_weekly_report():
    """Scheduled task for weekly report generation"""
    reporter = AutomatedReporter()
    try:
        reporter.generate_weekly_report()
    finally:
        reporter.close()


if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description="Automated Reporting System")
    parser.add_argument('--daily', action='store_true', help='Generate daily report')
    parser.add_argument('--weekly', action='store_true', help='Generate weekly report')
    parser.add_argument('--google-sheets', action='store_true', help='Export to Google Sheets')
    
    args = parser.parse_args()
    
    reporter = AutomatedReporter()
    
    try:
        if args.daily:
            reporter.generate_daily_report()
        
        if args.weekly:
            reporter.generate_weekly_report()
        
        if args.google_sheets:
            reporter.export_to_google_sheets()
        
        if not any([args.daily, args.weekly, args.google_sheets]):
            print("Usage: python automated_reporting.py [--daily] [--weekly] [--google-sheets]")
    
    finally:
        reporter.close()
